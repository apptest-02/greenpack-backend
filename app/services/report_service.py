"""
Greenpack Pro — Report Generation Service
ReportLab PDF QC reports + openpyxl Excel export + Windows print
"""
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from app.config import get_settings

log = logging.getLogger(__name__)
settings = get_settings()


def generate_pdf_report(
    job_id: str,
    config: dict,
    scores: dict,
    text_errors: list,
    color_result: dict,
    ssim_result: dict,
    barcode_result: list,
    annotated_path: Optional[Path] = None,
    ocr_timeout: bool = False,
) -> Optional[Path]:
    """Generate branded PDF QC report"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table,
            TableStyle, Image as RLImage, PageBreak, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

        # Output path
        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_report.pdf"

        # Colors - Professional palette
        NAVY = colors.HexColor("#0D1B2A")      # Dark navy for headers
        DARK_BLUE = colors.HexColor("#1A73E8")  # Primary blue
        TEAL = colors.HexColor("#00C2CB")       # Accent teal
        GREEN = colors.HexColor("#22A06B")      # Success green
        RED = colors.HexColor("#E5383B")        # Error red
        LIGHT_GRAY = colors.HexColor("#F5F7FA")  # Light background
        BORDER_GRAY = colors.HexColor("#E2E8F0")  # Border color
        TEXT_DARK = colors.HexColor("#1E293B")   # Dark text
        TEXT_MEDIUM = colors.HexColor("#475569") # Medium text
        TEXT_LIGHT = colors.HexColor("#94A3B8")  # Light text
        WHITE = colors.HexColor("#FFFFFF")       # White

        # Override with client brand color if configured
        brand_color = colors.HexColor(config.get("brand_color", "#0D1B2A"))

        # Styles
        def ps(name, **kw):
            d = dict(fontName="Helvetica", fontSize=10, leading=14)
            d.update(kw)
            return ParagraphStyle(name, **d)

        # Header styles with proper colors
        title_style = ps("title", fontName="Helvetica-Bold", fontSize=22,
                         textColor=WHITE, alignment=TA_CENTER, leading=28)
        h1_style = ps("h1", fontName="Helvetica-Bold", fontSize=14, 
                      textColor=NAVY, leading=18, spaceAfter=6)
        h2_style = ps("h2", fontName="Helvetica-Bold", fontSize=11, 
                      textColor=NAVY, leading=14)
        
        # Body styles
        body_style = ps("body", fontName="Helvetica", fontSize=9, 
                        textColor=TEXT_DARK, leading=12)
        bold_style = ps("bold", fontName="Helvetica-Bold", fontSize=9, 
                        textColor=NAVY, leading=12)
        small_style = ps("small", fontName="Helvetica", fontSize=7.5, 
                         textColor=TEXT_MEDIUM, leading=10)
        small_bold_style = ps("small_bold", fontName="Helvetica-Bold", fontSize=7.5, 
                               textColor=NAVY, leading=10)
        
        # Score styles
        score_style = ps("score", fontName="Helvetica-Bold", fontSize=52,
                         alignment=TA_CENTER, textColor=GREEN if scores["overall"] >= 75 else RED, leading=60)
        status_style = ps("status", fontName="Helvetica-Bold", fontSize=28,
                          alignment=TA_CENTER, leading=36)
        
        # Table header style
        th_style = ps("th", fontName="Helvetica-Bold", fontSize=8.5, 
                      textColor=WHITE, alignment=TA_CENTER, leading=12)

        # Page template with header/footer
        def page_template(canvas, doc):
            canvas.saveState()
            # Header bar
            canvas.setFillColor(brand_color)
            canvas.rect(0, A4[1] - 35, A4[0], 35, fill=1, stroke=0)
            canvas.setFont("Helvetica-Bold", 10)
            canvas.setFillColor(WHITE)
            canvas.drawString(1.5*cm, A4[1] - 23, "Greenpack Pro — Quality Control Report")
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(A4[0] - 1.5*cm, A4[1] - 23,
                f"Job: {config.get('job_ref', job_id[:8])}")
            # Footer
            canvas.setFillColor(LIGHT_GRAY)
            canvas.rect(0, 0, A4[0], 22, fill=1, stroke=0)
            canvas.setFont("Helvetica", 7)
            canvas.setFillColor(TEXT_LIGHT)
            canvas.drawString(1.5*cm, 7,
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Greenpack Pro v1.0")
            canvas.drawRightString(A4[0] - 1.5*cm, 7, f"Page {doc.page}")
            canvas.restoreState()

        # Build document
        doc = SimpleDocTemplate(
            str(out_path), pagesize=A4,
            leftMargin=1.8*cm, rightMargin=1.8*cm,
            topMargin=2.2*cm, bottomMargin=1.8*cm
        )

        story = []

        # ── Cover Section ──────────────────────────────────────────────────────
        overall = scores["overall"]
        status = "PASS" if overall >= 75 else "FAIL"
        status_color = GREEN if overall >= 75 else RED

        # Product title card
        cover_data = [[
            Paragraph(f'<font color="{WHITE.hexval()}"><b>{config.get("product_name", "Label Inspection Report")}</b></font>',
                      ps("cn", fontName="Helvetica-Bold", fontSize=18, textColor=WHITE, alignment=TA_CENTER)),
        ]]
        cover_tbl = Table(cover_data, colWidths=[A4[0] - 3.6*cm])
        cover_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), brand_color),
            ("TOPPADDING", (0, 0), (-1, -1), 15),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 15),
        ]))
        story.append(cover_tbl)
        story.append(Spacer(1, 15))

        # Score card
        score_data = [[
            Paragraph(f'<b>{overall:.1f}</b>', score_style),
            Paragraph(f'<font color="{status_color.hexval()}"><b>{status}</b></font>', status_style),
        ]]
        score_tbl = Table(score_data, colWidths=[7*cm, 10*cm])
        score_tbl.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 35),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 35),
            ("BACKGROUND", (0, 0), (-1, -1), WHITE),
            ("BOX", (0, 0), (-1, -1), 1.5, BORDER_GRAY),
        ]))
        story.append(score_tbl)
        story.append(Spacer(1, 15))

        # Sub-scores grid
        sub_data = [
            [Paragraph("OCR / Text", h2_style),
             Paragraph("Color", h2_style),
             Paragraph("Print Quality", h2_style),
             Paragraph("Barcode", h2_style)],
            [
                Paragraph(f'<font size=18 color="{GREEN.hexval() if scores["ocr"] >= 75 else RED.hexval()}"><b>{scores["ocr"]:.0f}</b></font>', ps("sv", alignment=TA_CENTER)),
                Paragraph(f'<font size=18 color="{GREEN.hexval() if scores["color"] >= 75 else RED.hexval()}"><b>{scores["color"]:.0f}</b></font>', ps("sv", alignment=TA_CENTER)),
                Paragraph(f'<font size=18 color="{GREEN.hexval() if scores["ssim"] >= 75 else RED.hexval()}"><b>{scores["ssim"]:.0f}</b></font>', ps("sv", alignment=TA_CENTER)),
                Paragraph(f'<font size=18 color="{GREEN.hexval() if scores["barcode"] >= 75 else RED.hexval()}"><b>{scores["barcode"]:.0f}</b></font>', ps("sv", alignment=TA_CENTER)),
            ],
        ]
        sub_tbl = Table(sub_data, colWidths=[(A4[0] - 3.6*cm) / 4] * 4)
        sub_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY),
            ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
            ("TOPPADDING", (0, 1), (-1, 1), 15),
            ("BOTTOMPADDING", (0, 1), (-1, 1), 15),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ]))
        story.append(sub_tbl)
        story.append(Spacer(1, 20))

        # Job details card
        story.append(Paragraph("Inspection Details", h1_style))
        story.append(HRFlowable(width="100%", thickness=1, color=TEAL, spaceBefore=5, spaceAfter=10))

        # Use simple text without HTML tags
        job_ref_val = config.get("job_ref", job_id[:12])
        client_val = config.get("client_name", "—")
        product_val = config.get("product_name", "—")
        inspector_val = config.get("inspector_name", "—")
        date_val = datetime.now().strftime("%Y-%m-%d %H:%M")
        time_val = f'{config.get("processing_time_ms", 0) / 1000:.1f}s'

        details_data = [
            [Paragraph("<b>Job Reference:</b>", bold_style), Paragraph(job_ref_val, body_style),
             Paragraph("<b>Client:</b>", bold_style), Paragraph(client_val, body_style)],
            [Paragraph("<b>Product:</b>", bold_style), Paragraph(product_val, body_style),
             Paragraph("<b>Inspector:</b>", bold_style), Paragraph(inspector_val, body_style)],
            [Paragraph("<b>Date:</b>", bold_style), Paragraph(date_val, body_style),
             Paragraph("<b>Processing Time:</b>", bold_style), Paragraph(time_val, body_style)],
        ]
        
        details_tbl = Table(details_data, colWidths=[3*cm, 7*cm, 2.5*cm, 4.5*cm])
        details_tbl.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GRAY),
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [LIGHT_GRAY, WHITE]),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, BORDER_GRAY),
        ]))
        story.append(details_tbl)
        story.append(Spacer(1, 20))

        # ── OCR Errors ─────────────────────────────────────────────────────────
        if text_errors:
            story.append(Paragraph("OCR / Text Errors", h1_style))
            story.append(HRFlowable(width="100%", thickness=1, color=TEAL, spaceBefore=5, spaceAfter=10))

            err_header = [["Region", "Master Text", "Scanned Text", "Type", "Severity"]]
            err_rows = []
            for e in text_errors[:20]:
                severity = e.get("severity", "")
                err_rows.append([
                    Paragraph(str(e.get("region_bbox", {}).get("x", "?"))[:10], small_style),
                    Paragraph(str(e.get("master_text", ""))[:35], small_style),
                    Paragraph(str(e.get("scan_text", ""))[:35], small_style),
                    Paragraph(e.get("type", ""), small_style),
                    Paragraph(severity.upper(), small_bold_style),
                ])

            err_tbl = Table(err_header + err_rows, colWidths=[1.8*cm, 5.5*cm, 5.5*cm, 2.5*cm, 2.2*cm])
            err_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT_GRAY, WHITE]),
                ("GRID", (0, 0), (-1, -1), 0.3, BORDER_GRAY),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(err_tbl)
            story.append(Spacer(1, 15))

        # ── Barcode Results ─────────────────────────────────────────────────────
        if barcode_result:
            story.append(Paragraph("Barcode Verification", h1_style))
            story.append(HRFlowable(width="100%", thickness=1, color=TEAL, spaceBefore=5, spaceAfter=10))

            bc_header = [["Type", "Decoded Value", "Expected", "Match", "Grade", "Status"]]
            bc_rows = []
            for b in barcode_result:
                status_txt = "PASS" if b.get("pass") else "FAIL"
                status_color_local = GREEN if b.get("pass") else RED
                bc_rows.append([
                    Paragraph(b.get("type", ""), small_style),
                    Paragraph(str(b.get("decoded_value", "N/A"))[:20], small_style),
                    Paragraph(str(b.get("expected_value", "—"))[:20], small_style),
                    Paragraph("Yes" if b.get("match") else "No", small_style),
                    Paragraph(str(b.get("quality_grade", "?")), small_style),
                    Paragraph(f'<font color="{status_color_local.hexval()}"><b>{status_txt}</b></font>', small_style),
                ])

            bc_tbl = Table(bc_header + bc_rows,
                          colWidths=[2*cm, 5*cm, 4.5*cm, 1.8*cm, 1.8*cm, 2.5*cm])
            bc_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT_GRAY, WHITE]),
                ("GRID", (0, 0), (-1, -1), 0.3, BORDER_GRAY),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(bc_tbl)
            story.append(Spacer(1, 15))

        # ── Color Results (simplified) ─────────────────────────────────────────
        if color_result.get("zone_results"):
            story.append(Paragraph("Color Analysis", h1_style))
            story.append(HRFlowable(width="100%", thickness=1, color=TEAL, spaceBefore=5, spaceAfter=10))
            
            col_data = [["Zone", "Name", "ΔE", "Status"]]
            for zone in color_result.get("zone_results", [])[:10]:
                col_data.append([
                    zone.get("zone_id", ""),
                    zone.get("zone_name", ""),
                    f'{zone.get("mean_delta_e", 0):.2f}',
                    "PASS" if zone.get("pass") else "FAIL",
                ])
            
            col_tbl = Table(col_data, colWidths=[2*cm, 4*cm, 2*cm, 2.5*cm])
            col_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT_GRAY, WHITE]),
                ("GRID", (0, 0), (-1, -1), 0.3, BORDER_GRAY),
            ]))
            story.append(col_tbl)
            story.append(Spacer(1, 15))

        # ── Annotated Image ─────────────────────────────────────────────────────
        if annotated_path and Path(annotated_path).exists():
            story.append(PageBreak())
            story.append(Paragraph("Annotated Label Comparison", h1_style))
            story.append(HRFlowable(width="100%", thickness=1, color=TEAL, spaceBefore=5, spaceAfter=10))
            img_width = A4[0] - 3.6*cm
            story.append(RLImage(str(annotated_path), width=img_width, height=img_width * 0.5))

        # Build PDF
        doc.build(story, onFirstPage=page_template, onLaterPages=page_template)
        log.info(f"PDF report generated: {out_path}")
        return out_path

    except Exception as e:
        log.error(f"PDF report generation failed: {e}")
        return None


def generate_excel_report(
    job_id: str,
    config: dict,
    scores: dict,
    text_errors: list,
    color_result: dict,
    ssim_result: dict,
    barcode_result: list,
) -> Optional[Path]:
    """Generate multi-sheet Excel QC report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        reports_dir = Path(settings.reports_dir)
        reports_dir.mkdir(parents=True, exist_ok=True)
        out_path = reports_dir / f"{job_id}_results.xlsx"

        wb = openpyxl.Workbook()

        # ── Summary Sheet ───────────────────────────────────────────────────────
        ws_sum = wb.active
        ws_sum.title = "Summary"

        # Define styles
        header_fill = PatternFill("solid", fgColor="0D1B2A")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        pass_fill = PatternFill("solid", fgColor="D1FAE5")
        fail_fill = PatternFill("solid", fgColor="FEE2E2")
        border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        summary_data = [
            ["Job Reference", config.get("job_ref", job_id[:12])],
            ["Client", config.get("client_name", "")],
            ["Product", config.get("product_name", "")],
            ["Date", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["", ""],
            ["Overall Score", f'{scores["overall"]:.1f}'],
            ["Status", "PASS" if scores["overall"] >= 75 else "FAIL"],
            ["", ""],
            ["OCR Score", f'{scores["ocr"]:.1f}'],
            ["Color Score", f'{scores["color"]:.1f}'],
            ["SSIM Score", f'{scores["ssim"]:.1f}'],
            ["Barcode Score", f'{scores["barcode"]:.1f}'],
            ["", ""],
            ["OCR Errors", len(text_errors)],
            ["Barcode Failures", sum(1 for b in barcode_result if not b.get("pass"))],
            ["Defects Found", len(ssim_result.get("defects", []))],
        ]

        for row_idx, (key, val) in enumerate(summary_data, 1):
            ws_sum.cell(row_idx, 1, key).font = Font(bold=True, size=10)
            ws_sum.cell(row_idx, 1).fill = PatternFill("solid", fgColor="F5F7FA")
            ws_sum.cell(row_idx, 2, str(val))
            ws_sum.cell(row_idx, 2).font = Font(size=10)
            if key == "Status":
                fill = pass_fill if val == "PASS" else fail_fill
                ws_sum.cell(row_idx, 2).fill = fill
                ws_sum.cell(row_idx, 2).font = Font(bold=True, size=11)
            if key == "Overall Score":
                ws_sum.cell(row_idx, 2).font = Font(bold=True, size=14)

        ws_sum.column_dimensions["A"].width = 22
        ws_sum.column_dimensions["B"].width = 30

        # ── OCR Errors Sheet ────────────────────────────────────────────────────
        if text_errors:
            ws_ocr = wb.create_sheet("OCR Errors")
            headers = ["Region X", "Region Y", "Master Text", "Scanned Text", "Type", "Severity"]
            for col, h in enumerate(headers, 1):
                cell = ws_ocr.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row_idx, err in enumerate(text_errors, 2):
                bbox = err.get("region_bbox", {})
                ws_ocr.cell(row_idx, 1, bbox.get("x", ""))
                ws_ocr.cell(row_idx, 2, bbox.get("y", ""))
                ws_ocr.cell(row_idx, 3, str(err.get("master_text", ""))[:100])
                ws_ocr.cell(row_idx, 4, str(err.get("scan_text", ""))[:100])
                ws_ocr.cell(row_idx, 5, err.get("type", ""))
                ws_ocr.cell(row_idx, 6, err.get("severity", ""))
                if err.get("severity") == "high":
                    for col in range(1, 7):
                        ws_ocr.cell(row_idx, col).fill = fail_fill
                for col in range(1, 7):
                    ws_ocr.cell(row_idx, col).border = border

            for col in range(1, 7):
                ws_ocr.column_dimensions[get_column_letter(col)].width = 20

        # ── Barcode Results Sheet ───────────────────────────────────────────────
        if barcode_result:
            ws_bc = wb.create_sheet("Barcode Results")
            bc_headers = ["Type", "Decoded Value", "Expected Value", "Match", "Check Digit", "Grade", "Status"]
            for col, h in enumerate(bc_headers, 1):
                cell = ws_bc.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row_idx, b in enumerate(barcode_result, 2):
                ws_bc.cell(row_idx, 1, b.get("type", ""))
                ws_bc.cell(row_idx, 2, str(b.get("decoded_value", "")))
                ws_bc.cell(row_idx, 3, str(b.get("expected_value", "")))
                ws_bc.cell(row_idx, 4, "Yes" if b.get("match") else "No")
                ws_bc.cell(row_idx, 5, "Yes" if b.get("check_digit_valid") else "No")
                ws_bc.cell(row_idx, 6, b.get("quality_grade", ""))
                status_val = "PASS" if b.get("pass") else "FAIL"
                ws_bc.cell(row_idx, 7, status_val)
                fill = pass_fill if b.get("pass") else fail_fill
                for col in range(1, 8):
                    ws_bc.cell(row_idx, col).fill = fill
                    ws_bc.cell(row_idx, col).border = border

            for col in range(1, 8):
                ws_bc.column_dimensions[get_column_letter(col)].width = 18

        # ── Color Results Sheet ─────────────────────────────────────────────────
        if color_result.get("zone_results"):
            ws_col = wb.create_sheet("Color Results")
            col_headers = ["Zone", "Zone Name", "Mean ΔE", "Max ΔE", "% Over", "Threshold", "Status"]
            for col, h in enumerate(col_headers, 1):
                cell = ws_col.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row_idx, zone in enumerate(color_result.get("zone_results", []), 2):
                ws_col.cell(row_idx, 1, zone.get("zone_id", ""))
                ws_col.cell(row_idx, 2, zone.get("zone_name", ""))
                ws_col.cell(row_idx, 3, zone.get("mean_delta_e", ""))
                ws_col.cell(row_idx, 4, zone.get("max_delta_e", ""))
                ws_col.cell(row_idx, 5, zone.get("pct_out_of_tolerance", ""))
                ws_col.cell(row_idx, 6, zone.get("threshold", ""))
                status_val = "PASS" if zone.get("pass") else "FAIL"
                ws_col.cell(row_idx, 7, status_val)
                fill = pass_fill if zone.get("pass") else fail_fill
                ws_col.cell(row_idx, 7).fill = fill
                for col in range(1, 8):
                    ws_col.cell(row_idx, col).border = border

            for col in range(1, 8):
                ws_col.column_dimensions[get_column_letter(col)].width = 14

        # ── Defects Sheet ───────────────────────────────────────────────────────
        if ssim_result.get("defects"):
            ws_def = wb.create_sheet("Defects")
            def_headers = ["Type", "Severity", "X", "Y", "Width", "Height", "Area (px²)"]
            for col, h in enumerate(def_headers, 1):
                cell = ws_def.cell(1, col, h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for row_idx, d in enumerate(ssim_result.get("defects", []), 2):
                bbox = d.get("bbox", {})
                ws_def.cell(row_idx, 1, d.get("type", ""))
                ws_def.cell(row_idx, 2, d.get("severity", ""))
                ws_def.cell(row_idx, 3, bbox.get("x", ""))
                ws_def.cell(row_idx, 4, bbox.get("y", ""))
                ws_def.cell(row_idx, 5, bbox.get("w", ""))
                ws_def.cell(row_idx, 6, bbox.get("h", ""))
                ws_def.cell(row_idx, 7, d.get("area_pixels", ""))
                if d.get("severity") == "critical":
                    for col in range(1, 8):
                        ws_def.cell(row_idx, col).fill = fail_fill
                for col in range(1, 8):
                    ws_def.cell(row_idx, col).border = border

            for col in range(1, 8):
                ws_def.column_dimensions[get_column_letter(col)].width = 14

        wb.save(str(out_path))
        log.info(f"Excel report generated: {out_path}")
        return out_path

    except Exception as e:
        log.error(f"Excel report generation failed: {e}")
        return None


def print_report_windows(report_path: str, printer_name: str = None) -> bool:
    """Send PDF report directly to Windows printer"""
    try:
        import win32api
        import win32con
        win32api.ShellExecute(
            0, "print", report_path, None, ".", 0
        )
        log.info(f"Report sent to Windows printer: {report_path}")
        return True
    except ImportError:
        # Fallback: open with default application
        import os
        os.startfile(report_path, "print")
        return True
    except Exception as e:
        log.error(f"Print failed: {e}")
        return False